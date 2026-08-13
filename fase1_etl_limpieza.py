#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script: fase1_etl_limpieza.py
Descripción: Script de ETL y limpieza de datos para procesar el archivo Excel
             'CONTROL EMPRESAS COPIA.xlsx' y generar archivos estandarizados CSV/JSON
             en el directorio /data_processed/.
"""

import os
import re
import json
import datetime
import warnings
import pandas as pd
import openpyxl

warnings.filterwarnings('ignore', category=UserWarning)

# -----------------------------------------------------------------------------
# Constantes y Configuración
# -----------------------------------------------------------------------------
EXCEL_FILE = 'CONTROL EMPRESAS COPIA.xlsx'
OUTPUT_DIR = 'data_processed'

REGIONAL_SHEETS = [
    'SOCIEDADES R.R--G.R',
    'SOCIEDADES R.G.R',
    'SOCIEDADES G.J.R',
    'AFAC-CAPI',
    'COLEGIOS ',
    'C-NORTE',
    'C-OAXACA '
]

# -----------------------------------------------------------------------------
# Funciones auxiliares de limpieza
# -----------------------------------------------------------------------------

def clean_str(val):
    """Limpia cadenas de texto: quita espacios borde, convierte nulos a None."""
    if val is None or pd.isna(val):
        return None
    s = str(val).strip()
    if not s or s.lower() in ['nan', 'none', 'null', 'n/a', 's/n', '']:
        return None
    return s

def clean_rfc(val):
    """Normaliza RFC: mayúsculas, sin espacios internos ni caracteres extraños."""
    s = clean_str(val)
    if not s:
        return None
    s = re.sub(r'\s+', '', s.upper())
    s = re.sub(r'[^A-Z0-9&Ñ]', '', s)
    return s if s else None

def is_valid_rfc(rfc):
    """Verifica si un RFC tiene el formato oficial mexicano (12 o 13 caracteres)."""
    if not rfc:
        return False
    pattern = r'^[A-Z&Ñ]{3,4}\d{6}[A-Z0-9]{3}$'
    return bool(re.match(pattern, rfc))

def parse_date(val):
    """
    Intenta parsear una fecha a formato ISO YYYY-MM-DD.
    Retorna tupla (fecha_iso, fecha_texto).
    """
    if val is None or pd.isna(val):
        return None, None

    if isinstance(val, (datetime.datetime, datetime.date, pd.Timestamp)):
        return val.strftime('%Y-%m-%d'), None

    s = str(val).strip()
    if not s or s.lower() in ['nan', 'none', 'null', 'n/a', 's/n', '']:
        return None, None

    try:
        dt = pd.to_datetime(s, dayfirst=True, errors='coerce')
        if pd.notna(dt) and 1900 <= dt.year <= 2100:
            return dt.strftime('%Y-%m-%d'), None
    except Exception:
        pass

    return None, s

# -----------------------------------------------------------------------------
# Carga optimizada de pestañas con openpyxl (read_only)
# -----------------------------------------------------------------------------

def load_sheet_rows(wb, sheet_name):
    """Carga las filas de una pestaña de forma eficiente excluyendo columnas vacías."""
    if sheet_name not in wb.sheetnames:
        return [], []

    ws = wb[sheet_name]
    rows_iter = ws.iter_rows(values_only=True)

    try:
        header_row = next(rows_iter)
    except StopIteration:
        return [], []

    # Determinar última columna con encabezado válido
    last_idx = len(header_row)
    while last_idx > 0 and (header_row[last_idx - 1] is None or str(header_row[last_idx - 1]).strip() == ''):
        last_idx -= 1

    headers = [str(header_row[i]).strip() if header_row[i] is not None else f'COL_{i}' for i in range(last_idx)]

    data_rows = []
    for r in rows_iter:
        row_slice = r[:last_idx]
        if any(v is not None and str(v).strip() != '' for v in row_slice):
            data_rows.append(row_slice)

    return headers, data_rows

# -----------------------------------------------------------------------------
# Procesadores de Pestañas Específicas
# -----------------------------------------------------------------------------

def process_control_empresas(headers, rows):
    """Procesa la pestaña principal 'CONTROL DE EMPRESAS'."""
    empresas_list = []
    socios_list = []

    col_idx = {h: i for i, h in enumerate(headers)}
    
    socio_indices = []
    for num in range(1, 7):
        key = f'SOCIO {num}'
        if key in col_idx:
            idx = col_idx[key]
            socio_indices.append((idx, idx + 1))

    for row in rows:
        rs_val = row[col_idx['RAZON SOCIAL']] if 'RAZON SOCIAL' in col_idx and col_idx['RAZON SOCIAL'] < len(row) else None
        rfc_val = row[col_idx['RFC']] if 'RFC' in col_idx and col_idx['RFC'] < len(row) else None

        razon_social = clean_str(rs_val)
        rfc = clean_rfc(rfc_val)

        if not razon_social and not rfc:
            continue

        fecha_val = row[col_idx['FECHA']] if 'FECHA' in col_idx and col_idx['FECHA'] < len(row) else None
        fecha_iso, fecha_txt = parse_date(fecha_val)

        def g(col_name):
            if col_name in col_idx and col_idx[col_name] < len(row):
                return clean_str(row[col_idx[col_name]])
            return None

        empresa = {
            'razon_social': razon_social,
            'rfc': rfc,
            'tipo_empresa': g('TIPO DE EMPRESA'),
            'numero_escritura': g('# ESCRITURA'),
            'rpp': g('RPP'),
            'fecha': fecha_iso,
            'fecha_texto': fecha_txt,
            'notaria': g('NOTARIA'),
            'domicilio_social': g('DOMICILIO SOCIAL'),
            'duracion': g('DURACION'),
            'capital_total_fijo': g('CAPITAL TOTAL FIJO'),
            'administrador_unico_gerente': g('ADMINISTRADOR UNICO/GERENTE'),
            'apoderados': g('APODERADOS'),
            'comisario': g('COMISARIO'),
            'delegado': g('DELEGADO'),
            'asa_venta': g('ASA VENTA'),
            'numero_poder_revocacion': g('#PODER/ REVOCACIÓN'),
            'modificacion_estatutos': g('MODIFICACIÓN DE STATUTOS'),
            'afac_capi': g('AFAC-CAPI'),
            'observacion': g('OBSERVACION'),
            'origen_tags': ['CONTROL DE EMPRESAS']
        }
        empresas_list.append(empresa)

        # Desnormalizar socios
        for name_idx, pct_idx in socio_indices:
            socio_name = clean_str(row[name_idx]) if name_idx < len(row) else None
            if socio_name:
                pct_val = row[pct_idx] if pct_idx < len(row) else None
                socios_list.append({
                    'rfc_empresa': rfc,
                    'razon_social_empresa': razon_social,
                    'nombre_socio': socio_name,
                    'porcentaje_participacion': clean_str(pct_val),
                    'tipo_socio': 'CAPITAL_FIJO',
                    'origen_tabla': 'CONTROL DE EMPRESAS'
                })

    return empresas_list, socios_list

def process_domicilio(headers, rows):
    """Procesa la pestaña 'DOMICILIO'."""
    domicilios = []
    col_idx = {h: i for i, h in enumerate(headers)}

    for row in rows:
        def g(col_name):
            if col_name in col_idx and col_idx[col_name] < len(row):
                return clean_str(row[col_idx[col_name]])
            return None

        razon_social = g('RAZON SOCIAL')
        rfc = clean_rfc(g('RFC'))

        if not razon_social and not rfc:
            continue

        domicilios.append({
            'razon_social': razon_social,
            'rfc': rfc,
            'estado': g('ESTADO'),
            'municipio_delegacion': g('MUNICIPIO O DELEGACION'),
            'conocido': g('CONOCIDO ') or g('CONOCIDO'),
            'domicilio_fiscal': g('DOMICILIO FISCAL'),
            'estatus': g('ESTATUS')
        })
    return domicilios

def process_ventas(headers, rows):
    """Procesa la pestaña 'VENTAS'."""
    ventas = []
    extra_socios = []
    col_idx = {h: i for i, h in enumerate(headers)}

    socio_indices = []
    for num in range(1, 7):
        key = f'SOCIO {num}'
        if key in col_idx:
            idx = col_idx[key]
            socio_indices.append((idx, idx + 1))

    for row in rows:
        def g(col_name):
            if col_name in col_idx and col_idx[col_name] < len(row):
                return clean_str(row[col_idx[col_name]])
            return None

        razon_social = g('DENOMINACION SOCIAL') or g('RAZON SOCIAL')
        rfc = clean_rfc(g('RFC'))

        if not razon_social and not rfc:
            continue

        fecha_val = row[col_idx['FECHA']] if 'FECHA' in col_idx and col_idx['FECHA'] < len(row) else None
        fecha_iso, fecha_txt = parse_date(fecha_val)

        ventas.append({
            'razon_social': razon_social,
            'rfc': rfc,
            'tipo_empresa': g('TIPO DE EMPRESA'),
            'numero_escritura': g('# ESCRITURA'),
            'rpp': g('RPP'),
            'fecha': fecha_iso,
            'fecha_texto': fecha_txt,
            'notaria': g('NOTARIA'),
            'documento': g('DOCUMENTO'),
            'domicilio_social': g('DOMICILIO SOCIAL'),
            'capital_total_fijo': g('CAPITAL TOTAL FIJO'),
            'socios_capital_variable': g('SOCIOS DE CAPITAL VARIABLE ') or g('SOCIOS DE CAPITAL VARIABLE'),
            'administrador_unico_gerente': g('ADMINISTRADOR UNICO/GERENTE'),
            'apoderados': g('APODERADOS'),
            'comisario': g('COMISARIO'),
            'delegado': g('DELEGADO'),
            'escrutador': g('ESCRUTADOR ') or g('ESCRUTADOR'),
            'observaciones': g('OBSERVACIONES ') or g('OBSERVACIONES')
        })

        for name_idx, pct_idx in socio_indices:
            socio_name = clean_str(row[name_idx]) if name_idx < len(row) else None
            if socio_name:
                pct_val = row[pct_idx] if pct_idx < len(row) else None
                extra_socios.append({
                    'rfc_empresa': rfc,
                    'razon_social_empresa': razon_social,
                    'nombre_socio': socio_name,
                    'porcentaje_participacion': clean_str(pct_val),
                    'tipo_socio': 'VENTAS_SOCIO',
                    'origen_tabla': 'VENTAS'
                })

    return ventas, extra_socios

def process_poderes(headers, rows):
    """Procesa la pestaña 'PODERES-REVOCACIÓN'."""
    poderes = []
    col_idx = {h: i for i, h in enumerate(headers)}

    for row in rows:
        def g(col_name):
            if col_name in col_idx and col_idx[col_name] < len(row):
                return clean_str(row[col_idx[col_name]])
            return None

        razon_social = g('RAZON SOCIAL')
        rfc = clean_rfc(g('RFC'))

        if not razon_social and not rfc:
            continue

        fecha_val = row[col_idx['FECHA']] if 'FECHA' in col_idx and col_idx['FECHA'] < len(row) else None
        fecha_iso, fecha_txt = parse_date(fecha_val)

        poderes.append({
            'razon_social': razon_social,
            'rfc': rfc,
            'tipo_empresa': g('TIPO DE EMPRESA'),
            'numero_escritura': g('# ESCRITURA'),
            'rpp': g('RPP'),
            'fecha': fecha_iso,
            'fecha_texto': fecha_txt,
            'notaria': g('NOTARIA'),
            'documento': g('DOCUMENTO'),
            'administrador_unico_gerente': g('ADMINISTRADOR UNICO/GERENTE'),
            'apoderados': g('APODERADOS'),
            'delegado': g('DELEGADO'),
            'observaciones': g('OBSERVACIONES')
        })
    return poderes

def process_modificacion_estatutos(headers, rows):
    """Procesa la pestaña 'ASA MODIFICACIÓN STATUTOS'."""
    modificaciones = []
    extra_socios = []
    col_idx = {h: i for i, h in enumerate(headers)}

    socio_indices = []
    for num in range(1, 6):
        key = f'SOCIO {num}'
        if key in col_idx:
            idx = col_idx[key]
            socio_indices.append((idx, idx + 1))

    for row in rows:
        def g(col_name):
            if col_name in col_idx and col_idx[col_name] < len(row):
                return clean_str(row[col_idx[col_name]])
            return None

        razon_social = g('RAZON SOCIAL')
        rfc = clean_rfc(g('RFC'))

        if not razon_social and not rfc:
            continue

        fecha_val = row[col_idx['FECHA']] if 'FECHA' in col_idx and col_idx['FECHA'] < len(row) else None
        fecha_iso, fecha_txt = parse_date(fecha_val)

        modificaciones.append({
            'razon_social': razon_social,
            'rfc': rfc,
            'numero_escritura': g('# ESCRITURA'),
            'rpp': g('RPP'),
            'fecha': fecha_iso,
            'fecha_texto': fecha_txt,
            'notaria': g('NOTARIA+') or g('NOTARIA'),
            'documento': g('DOCUMENTO'),
            'domicilio_social': g('DOMICILIO SOCIAL'),
            'capital_total_fijo': g('CAPITAL TOTAL FIJO'),
            'administrador_unico_gerente': g('ADMINISTRADOR UNICO/GERENTE'),
            'apoderados': g('APODERADOS'),
            'comisario': g('COMISARIO'),
            'delegado': g('DELEGADO'),
            'escrutador': g('ESCRUTADOR ') or g('ESCRUTADOR'),
            'observaciones': g('OBSERVACIONES')
        })

        for name_idx, pct_idx in socio_indices:
            socio_name = clean_str(row[name_idx]) if name_idx < len(row) else None
            if socio_name:
                pct_val = row[pct_idx] if pct_idx < len(row) else None
                extra_socios.append({
                    'rfc_empresa': rfc,
                    'razon_social_empresa': razon_social,
                    'nombre_socio': socio_name,
                    'porcentaje_participacion': clean_str(pct_val),
                    'tipo_socio': 'MODIFICACION_ESTATUTOS',
                    'origen_tabla': 'ASA MODIFICACIÓN STATUTOS'
                })

    return modificaciones, extra_socios

def process_regional_sheet(sheet_name, headers, rows):
    """Procesa una pestaña regional e identifica empresas y socios."""
    companies = []
    socios = []

    col_idx = {h: i for i, h in enumerate(headers)}

    socio_indices = []
    for num in range(1, 7):
        key = f'SOCIO {num}'
        if key in col_idx:
            idx = col_idx[key]
            socio_indices.append((idx, idx + 1))

    rs_key = next((h for h in ['RAZÓN SOCIAL', 'RAZON SOCIAL'] if h in col_idx), None)

    for row in rows:
        def g(col_name):
            if col_name in col_idx and col_idx[col_name] < len(row):
                return clean_str(row[col_idx[col_name]])
            return None

        razon_social = g(rs_key) if rs_key else None
        rfc = clean_rfc(g('RFC'))

        if not razon_social and not rfc:
            continue

        fecha_raw = g('FECHA') or g('FECHA PROTOCOLIZACIÓN ')
        fecha_iso, fecha_txt = parse_date(fecha_raw)

        comp = {
            'razon_social': razon_social,
            'rfc': rfc,
            'tipo_empresa': g('TIPO DE EMPRESA'),
            'numero_escritura': g('# ESCRITURA'),
            'rpp': g('RPP'),
            'fecha': fecha_iso,
            'fecha_texto': fecha_txt,
            'notaria': g('NOTARIA'),
            'domicilio_social': g('DOMICILIO SOCIAL'),
            'duracion': g('DURACION DE LA SOCIEDAD') or g('DURACION'),
            'capital_total_fijo': g('CAPITAL TOTAL FIJO') or g('CAPITALTOTAL'),
            'administrador_unico_gerente': g('ADMINISTRADOR UNICO/GERENTE'),
            'apoderados': g('APODERADOS'),
            'comisario': g('COMISARIO'),
            'delegado': g('DELEGADO'),
            'asa_venta': None,
            'numero_poder_revocacion': g('PODER'),
            'modificacion_estatutos': None,
            'afac_capi': g('FECHA AFAC/CAPI') or g('APORTACIÓN DE AFAC'),
            'observacion': g('OBSERVACIONES') or g('OBSERVACIONES '),
            'origen_tags': [sheet_name.strip()]
        }
        companies.append(comp)

        for name_idx, pct_idx in socio_indices:
            socio_name = clean_str(row[name_idx]) if name_idx < len(row) else None
            if socio_name:
                pct_val = row[pct_idx] if pct_idx < len(row) else None
                socios.append({
                    'rfc_empresa': rfc,
                    'razon_social_empresa': razon_social,
                    'nombre_socio': socio_name,
                    'porcentaje_participacion': clean_str(pct_val),
                    'tipo_socio': f'REGIONAL_{sheet_name.strip()}',
                    'origen_tabla': sheet_name.strip()
                })

    return companies, socios

# -----------------------------------------------------------------------------
# Función Principal Execution
# -----------------------------------------------------------------------------

def main():
    print("=== INICIANDO FASE 1: ETL Y LIMPIEZA DE DATOS ===", flush=True)

    if not os.path.exists(EXCEL_FILE):
        raise FileNotFoundError(f"No se encontró el archivo de origen: {EXCEL_FILE}")

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True, data_only=True)

    audit_rows_read = {}

    # 1. Pestaña Principal: CONTROL DE EMPRESAS
    headers_ce, rows_ce = load_sheet_rows(wb, 'CONTROL DE EMPRESAS')
    audit_rows_read['CONTROL DE EMPRESAS'] = len(rows_ce)
    main_empresas, main_socios = process_control_empresas(headers_ce, rows_ce)

    # Estructura de consolidación inteligente
    master_companies = []
    rfc_map = {}
    name_map = {}

    def normalize_name(name):
        if not name:
            return None
        return re.sub(r'\s+', ' ', name.strip().upper())

    def add_or_merge_company(emp):
        rfc = emp.get('rfc')
        rs = emp.get('razon_social')
        norm_rs = normalize_name(rs)

        existing = None
        if rfc and rfc in rfc_map:
            existing = rfc_map[rfc]
        elif norm_rs and norm_rs in name_map:
            existing = name_map[norm_rs]

        if existing:
            # Actualizar RFC o Razón Social si antes eran NULL
            if not existing['rfc'] and rfc:
                existing['rfc'] = rfc
                rfc_map[rfc] = existing
            if not existing['razon_social'] and rs:
                existing['razon_social'] = rs
                if norm_rs:
                    name_map[norm_rs] = existing

            # Fusionar tags
            for tag in emp.get('origen_tags', []):
                if tag not in existing['origen_tags']:
                    existing['origen_tags'].append(tag)

            # Rellenar campos faltantes
            for col, val in emp.items():
                if col != 'origen_tags' and val is not None and existing.get(col) is None:
                    existing[col] = val
        else:
            # Nueva empresa en la lista maestra
            master_companies.append(emp)
            if rfc:
                rfc_map[rfc] = emp
            if norm_rs:
                name_map[norm_rs] = emp

    # Insertar empresas principales
    for emp in main_empresas:
        add_or_merge_company(emp)

    all_socios = list(main_socios)

    # 2. Pestañas Regionales
    for r_sheet in REGIONAL_SHEETS:
        headers_r, rows_r = load_sheet_rows(wb, r_sheet)
        audit_rows_read[r_sheet.strip()] = len(rows_r)
        r_companies, r_socios = process_regional_sheet(r_sheet, headers_r, rows_r)
        
        all_socios.extend(r_socios)

        for emp in r_companies:
            add_or_merge_company(emp)

    # 3. Pestaña DOMICILIO
    headers_dom, rows_dom = load_sheet_rows(wb, 'DOMICILIO')
    audit_rows_read['DOMICILIO'] = len(rows_dom)
    domicilios_list = process_domicilio(headers_dom, rows_dom)

    # 4. Pestaña VENTAS
    headers_vta, rows_vta = load_sheet_rows(wb, 'VENTAS')
    audit_rows_read['VENTAS'] = len(rows_vta)
    ventas_list, vta_socios = process_ventas(headers_vta, rows_vta)
    all_socios.extend(vta_socios)

    # 5. Pestaña PODERES-REVOCACIÓN
    headers_pod, rows_pod = load_sheet_rows(wb, 'PODERES-REVOCACIÓN')
    audit_rows_read['PODERES-REVOCACIÓN'] = len(rows_pod)
    poderes_list = process_poderes(headers_pod, rows_pod)

    # 6. Pestaña ASA MODIFICACIÓN STATUTOS
    headers_mod, rows_mod = load_sheet_rows(wb, 'ASA MODIFICACIÓN STATUTOS')
    audit_rows_read['ASA MODIFICACIÓN STATUTOS'] = len(rows_mod)
    modificaciones_list, mod_socios = process_modificacion_estatutos(headers_mod, rows_mod)
    all_socios.extend(mod_socios)

    wb.close()

    # Consolidar DataFrames
    df_empresas = pd.DataFrame(master_companies)
    df_socios = pd.DataFrame(all_socios)
    df_domicilios = pd.DataFrame(domicilios_list)
    df_ventas = pd.DataFrame(ventas_list)
    df_poderes = pd.DataFrame(poderes_list)
    df_modificaciones = pd.DataFrame(modificaciones_list)

    # Convertir origen_tags a cadena separada por comas para CSV
    df_empresas_csv = df_empresas.copy()
    if 'origen_tags' in df_empresas_csv.columns:
        df_empresas_csv['origen_tags'] = df_empresas_csv['origen_tags'].apply(
            lambda x: ", ".join(x) if isinstance(x, list) else str(x)
        )

    # -------------------------------------------------------------------------
    # Auditoría de RFCs
    # -------------------------------------------------------------------------
    missing_rfc_count = 0
    invalid_rfc_count = 0

    for emp in master_companies:
        rfc = emp.get('rfc')
        if not rfc:
            missing_rfc_count += 1
        elif not is_valid_rfc(rfc):
            invalid_rfc_count += 1

    # -------------------------------------------------------------------------
    # Guardar Salidas (JSON + CSV)
    # -------------------------------------------------------------------------
    outputs = [
        ('empresas', df_empresas, df_empresas_csv),
        ('socios', df_socios, df_socios),
        ('domicilios', df_domicilios, df_domicilios),
        ('ventas', df_ventas, df_ventas),
        ('poderes_revocacion', df_poderes, df_poderes),
        ('modificacion_estatutos', df_modificaciones, df_modificaciones)
    ]

    print("\n--- EXPORTANDO ARCHIVOS ---", flush=True)
    for name, df_json, df_csv in outputs:
        json_path = os.path.join(OUTPUT_DIR, f"{name}.json")
        csv_path = os.path.join(OUTPUT_DIR, f"{name}.csv")

        # Exportar JSON
        df_json.to_json(json_path, orient='records', force_ascii=False, indent=2)
        
        # Exportar CSV
        df_csv.to_csv(csv_path, index=False, encoding='utf-8-sig')

        print(f"  [+] Generados: {json_path} ({len(df_json)} registros) y {csv_path}")

    # -------------------------------------------------------------------------
    # Generar Reporte de Auditoría
    # -------------------------------------------------------------------------
    report_path = os.path.join(OUTPUT_DIR, 'reporte_auditoria.txt')
    report_lines = [
        "=" * 60,
        "REPORTE DE AUDITORÍA Y CONSOLIDACIÓN DE DATOS",
        f"Fecha de Ejecución: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        "=" * 60,
        "",
        "1. RESUMEN DE FILAS LEÍDAS POR PESTAÑA:"
    ]
    for sheet_name, count in audit_rows_read.items():
        report_lines.append(f"   - {sheet_name:30s}: {count:5d} filas")

    report_lines.extend([
        "",
        "2. RESUMEN DE TABLAS CONSOLIDADAS Y REGISTROS EXPORTADOS:",
        f"   - Empresas Únicas Consolidadas : {len(df_empresas):5d} registros",
        f"   - Socios Extraídos               : {len(df_socios):5d} registros",
        f"   - Domicilios                     : {len(df_domicilios):5d} registros",
        f"   - Ventas                         : {len(df_ventas):5d} registros",
        f"   - Poderes y Revocaciones         : {len(df_poderes):5d} registros",
        f"   - Modificaciones de Estatutos    : {len(df_modificaciones):5d} registros",
        "",
        "3. AUDITORÍA DE RFC EN EMPRESAS CONSOLIDADAS:",
        f"   - Total Empresas                 : {len(df_empresas):5d}",
        f"   - Empresas sin RFC (NULL / Vacío): {missing_rfc_count:5d}",
        f"   - Empresas con RFC Inválido      : {invalid_rfc_count:5d}",
        f"   - Empresas con RFC Válido        : {len(df_empresas) - missing_rfc_count - invalid_rfc_count:5d}",
        "",
        "=" * 60
    ])

    report_content = "\n".join(report_lines)
    with open(report_path, 'w', encoding='utf-8') as f:
        f.write(report_content)

    print("\n--- REPORTE DE AUDITORÍA GENERADO ---", flush=True)
    print(report_content)
    print("=== PROCESO FINALIZADO EXITOSAMENTE ===", flush=True)

if __name__ == '__main__':
    main()
